import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


TRACKER_PATH = "data/job_tracker.xlsx"

COLUMNS = [
    "date_found",
    "campaign",
    "job_title",
    "company",
    "source",
    "location",
    "match_score",
    "job_url",
    "cover_letter_path",
    "ats_answers_path",
    "status",
    "date_applied",
    "interview_date",
    "notes",
]

# Display-friendly header names
HEADER_NAMES = [
    "Date Found",
    "Campaign",
    "Job Title",
    "Company",
    "Source",
    "Location",
    "Match Score",
    "Job URL",
    "Cover Letter",
    "ATS Answers",
    "Status",
    "Date Applied",
    "Interview Date",
    "Notes",
]

COLUMN_WIDTHS = {
    "Date Found":      14,
    "Campaign":        28,
    "Job Title":       35,
    "Company":         25,
    "Source":           14,
    "Location":        18,
    "Match Score":      12,
    "Job URL":         50,
    "Cover Letter":    40,
    "ATS Answers":     40,
    "Status":          12,
    "Date Applied":    14,
    "Interview Date":  14,
    "Notes":           40,
}

TABLE_NAME = "JobTracker"


def create_tracker(path: str = TRACKER_PATH):
    """Create a fresh Excel tracker with a proper Excel Table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    # Write headers
    for col_idx, header in enumerate(HEADER_NAMES, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value     = header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    for col_idx, header in enumerate(HEADER_NAMES, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 15)

    # Create Excel Table with just the header row
    # Table range starts at A1 and covers all columns
    last_col = get_column_letter(len(HEADER_NAMES))
    table_ref = f"A1:{last_col}1"

    table = Table(displayName=TABLE_NAME, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Freeze header row
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    wb.save(path)
    print(f"  Tracker created: {path}")


def load_existing_urls(path: str = TRACKER_PATH) -> set:
    """Return set of URLs already in the tracker."""
    if not os.path.exists(path):
        return set()
    try:
        df = pd.read_excel(path, sheet_name="Job Applications")
        for col_name in ["Job URL", "job_url", "Job Url"]:
            if col_name in df.columns:
                return set(df[col_name].dropna().tolist())
    except Exception:
        pass
    return set()


def expand_table_range(ws, new_last_row: int):
    """
    Update the Excel Table's ref to include new rows.
    This keeps the Table intact when rows are appended.
    """
    for table in ws.tables.values():
        if table.displayName == TABLE_NAME:
            last_col = get_column_letter(len(HEADER_NAMES))
            table.ref = f"A1:{last_col}{new_last_row}"
            return


def append_jobs(df: pd.DataFrame, campaign_name: str,
                path: str = TRACKER_PATH):
    """Append new jobs to the Excel tracker Table."""
    if df.empty:
        print(f"  No jobs to append for [{campaign_name}]")
        return 0

    if not os.path.exists(path):
        create_tracker(path)

    existing_urls = load_existing_urls(path)
    wb = load_workbook(path)
    ws = wb["Job Applications"]

    next_row = ws.max_row + 1
    added    = 0
    skipped  = 0

    for _, row in df.iterrows():
        url = str(row.get("url", ""))

        if url in existing_urls:
            skipped += 1
            continue

        row_data = [
            datetime.now().strftime("%Y-%m-%d"),           # Date Found
            campaign_name,                                  # Campaign
            str(row.get("title", "")),                     # Job Title
            str(row.get("company", "")),                   # Company
            str(row.get("source", "")),                    # Source
            str(row.get("location", "")),                  # Location
            round(float(row.get("match_score", 0)), 3),    # Match Score
            url,                                            # Job URL
            str(row.get("cover_letter_path", "")),         # Cover Letter
            str(row.get("ats_answers_path", "")),          # ATS Answers
            "New",                                          # Status
            "",                                             # Date Applied
            "",                                             # Interview Date
            "",                                             # Notes
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell       = ws.cell(row=next_row, column=col_idx)
            cell.value = value

            # Format match score
            if col_idx == 7:
                cell.number_format = "0.000"
                cell.alignment = Alignment(horizontal="center")

            # Center dates and status
            if col_idx in (1, 11, 12, 13):
                cell.alignment = Alignment(horizontal="center")

        existing_urls.add(url)
        next_row += 1
        added    += 1

    # Expand the Table range to include new rows
    if added > 0:
        expand_table_range(ws, next_row - 1)

    wb.save(path)

    if skipped > 0:
        print(f"  [{campaign_name}] {added} added, {skipped} already tracked")
    else:
        print(f"  [{campaign_name}] {added} rows added")

    return added


def run_tracker(tailored_results: dict,
                path: str = TRACKER_PATH) -> int:
    """
    Write all results to the Excel tracker.
    Input:  dict of {campaign_name: df_with_file_paths}
    Output: total rows added
    """
    print("\n=== MODULE 5: Excel tracker ===")

    if not os.path.exists(path):
        create_tracker(path)
        print("  First run — tracker created")

    total_added = 0
    for campaign_name, df in tailored_results.items():
        total_added += append_jobs(df, campaign_name, path)

    print(f"\n  Total new rows: {total_added}")
    print(f"  File: {os.path.abspath(path)}")
    print(f"\n=== Module 5 complete ===")
    return total_added


if __name__ == "__main__":
    from modules.scraper import run_scraper
    from modules.cleaner import run_cleaner
    from modules.matcher import run_matcher
    from modules.tailor  import run_tailor

    scraped  = run_scraper()
    cleaned  = run_cleaner(scraped)
    scored   = run_matcher(cleaned)
    tailored = run_tailor(scored)
    total    = run_tracker(tailored)

    print(f"\n--- Tracker test ---")
    print(f"  Rows written: {total}")

    if os.path.exists(TRACKER_PATH):
        df_check = pd.read_excel(TRACKER_PATH, sheet_name="Job Applications")
        print(f"  Total rows in file: {len(df_check)}")
        if not df_check.empty:
            print(f"  Latest: {df_check.iloc[-1].get('Job Title', '')} @ {df_check.iloc[-1].get('Company', '')}")

    print(f"\n✓ Module 5 test complete")